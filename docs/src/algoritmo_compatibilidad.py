#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
algoritmo_compatibilidad.py  

Algoritmo de emparejamiento por PORCENTAJE DE AFINIDAD para co-housing senior.


USO
  python algoritmo_compatibilidad.py              # todos los residentes
  python algoritmo_compatibilidad.py BMW7812      # sólo ese residente en consola
  python algoritmo_compatibilidad.py --top 5      # top-5 por residente
  python algoritmo_compatibilidad.py --sin-excel  # salta la generación Excel
"""

from __future__ import annotations

import sys, io, argparse, statistics
from pathlib import Path
from rdflib import Graph, URIRef, RDF
import pandas as pd


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
else:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")


# RUTAS
# Carpeta donde está este script
_HERE = Path(__file__).parent

# Fichero TTL de la ontología
TTL_FILE = _HERE / "co-Housing-extendida.ttl"

# Buscar el Excel con hobby y religión
XLS_FILE = None
for _ruta in [
    _HERE / "ResidenciaDB-xlsx.xls",
    Path.home() / "Downloads" / "ResidenciaDB-xlsx.xls",
]:
    if _ruta.exists():
        XLS_FILE = _ruta
        break

RESULTADOS = _HERE / "resultados"

H = "http://www.semanticweb.org/alvar/ontologies/2026/2/untitled-ontology-2/"
D = "http://datos.residencia/"


# TABLA DE PESOS  (todos los campos de la ontología)

PESO: dict[str, int] = {
    "mismo_hobby"             : +8,   
    "misma_religion"          : +6,   
    "religion_opuesta"        : -4,   
    "casado_con"              : +20,
    "es_familiar"             : +15,
    "mismo_idioma"            : +15,
    "mismo_ciclo_sueno"       : +12,
    "horas_sueno_similares"   : +4,
    "ciclo_opuesto"           : -14,
    "mismo_nivel_ruido"       : +10,
    "ruido_incompatible"      : -12,
    "ruido_tol_diferente"     : -5,
    "dependencia_igual"       : +10,
    "dependencia_1nivel"      : +3,
    "dependencia_2nivel"      : -6,
    "dependencia_3nivel"      : -12,
    "cognitivo_igual"         : +6,
    "cognitivo_2nivel"        : -6,
    "ambos_bajo_riesgo_caida" : +3,
    "ambos_alto_riesgo_caida" : -5,
    "adl_igual"               : +5,
    "adl_muy_diferente"       : -5,
    "ambos_asistencia_noc"    : +3,
    "asistencia_noc_dif"      : -3,
    "ambos_diabeticos"        : +4,
    "ambos_no_alcohol"        : +3,
    "perdida_memoria_doble"   : -5,
    "necesidad_medica_comun"  : +3,
    "afinidad_medica_comun"   : +2,
    "mismo_nivel_actividad"   : +7,
    "actividad_fisica_sim"    : +4,
    "actividad_muy_diferente" : -5,
    "misma_zona_comun"        : +6,
    "ambos_hab_compartida"    : +10,
    "genero_compatible"       : +5,
    "hab_discordante"         : -5,
    "misma_frec_visitas"      : +3,
    "mismo_vivio_solo"        : +3,
    "misma_dieta"             : +5,
    "mascota_afinidad"        : +5,
    "presupuesto_similar"     : +3,
    "presupuesto_muy_dif"     : -4,
    "estres_similar"          : +4,
    "estres_muy_diferente"    : -4,
    "edad_muy_diferente"      : -4,
    "ansiedad_doble"          : -8,
    "depresion_doble"         : -8,
}

ORDEN_RUIDO = {"Silencioso": 0, "Bajo": 1, "Medio": 2, "ActividadSocial": 3, "Alto": 4}
GENERO_MAP  = {"Mujer": "Female", "Hombre": "Male", "Varón": "Male"}


# FUNCIONES DE INTERPRETACIÓN CLÍNICA

def nivel_barthel(b):
    if b is None: return None
    if b >= 91: return 0
    if b >= 61: return 1
    if b >= 41: return 2
    if b >= 21: return 3
    return 4

ETIQ_BARTHEL = {0: "Independiente", 1: "Dep. leve", 2: "Dep. moderada",
                3: "Dep. severa",   4: "Dep. total"}

def nivel_meclobo(m):
    if m is None: return None
    if m >= 24: return 0
    if m >= 18: return 1
    if m >= 11: return 2
    return 3

ETIQ_MECLOBO = {0: "Normal", 1: "Det. leve", 2: "Det. mod.", 3: "Det. grave"}

def riesgo_tinetti(t):
    if t is None: return None
    if t > 18:  return "Bajo"
    if t >= 10: return "Medio"
    return "Alto"

def interp_gds(n):
    if n is None: return "—"
    if n <= 4:  return "Normal"
    if n <= 8:  return "Dep. leve"
    if n <= 11: return "Dep. moderada"
    return "Dep. grave"

def interp_goldberg(n):
    if n is None: return "—"
    if n <= 3: return "Sin riesgo"
    if n <= 5: return "Riesgo leve"
    if n <= 8: return "Riesgo mod."
    return "Riesgo alto"

def interp_norton(n):
    if n is None: return "—"
    if n <= 9:  return "Muy alto"
    if n <= 12: return "Alto"
    if n <= 14: return "Medio"
    return "Bajo"

def interp_npi(n):
    if n is None: return "—"
    if n == 0:  return "Sin alterac."
    if n <= 3:  return "Leve"
    if n <= 7:  return "Moderado"
    return "Grave"

def interp_pfeiffer(n):
    if n is None: return "—"
    if n <= 2:  return "Normal"
    if n <= 4:  return "Det. leve"
    if n <= 7:  return "Det. mod."
    return "Det. grave"

def interp_necpal(n):
    if n is None: return "—"
    return "Negativo" if n == 0 else "Positivo"

def n_adl(p: dict) -> int:
    return sum(1 for k in ("ayudaCaminar", "ayudaComer", "ayudaVestir")
               if p.get(k) is True)

# NIVEL DE CONVIVENCIA (replica las 6 reglas SWRL de la ontología)

#  Nivel 1 – Alteraciones de Conducta (Prioridad Clínica)  
#  Nivel 2 – Dependencia Total                             
#  Nivel 3 – Dependencia Severa                           
#  Nivel 4 – Dep. Moderada / Riesgo de Caídas              
#  Nivel 5 – Dep. Leve / Fragilidad Psicogeriátrica        
#  Nivel 6 – Independencia Total                           

NIVEL_LABEL = {
    1: "Nivel 1 – Alteraciones de Conducta",
    2: "Nivel 2 – Dependencia Total",
    3: "Nivel 3 – Dependencia Severa",
    4: "Nivel 4 – Dep. Moderada / Riesgo Caídas",
    5: "Nivel 5 – Dep. Leve / Fragilidad",
    6: "Nivel 6 – Independencia Total",
}
NIVEL_SHORT = {
    1: "N1 · Alt.Conducta",  2: "N2 · Dep.Total",
    3: "N3 · Dep.Severa",    4: "N4 · Dep.Moderada",
    5: "N5 · Dep.Leve",      6: "N6 · Independiente",
}

NIVEL_COLOR_BG = {
    1: "C62828", 2: "E64A19", 3: "F57C00",
    4: "F9A825", 5: "8BC34A", 6: "2E7D32",
}
NIVEL_COLOR_TXT = {
    1: "FFFFFF", 2: "FFFFFF", 3: "FFFFFF",
    4: "000000", 5: "000000", 6: "FFFFFF",
}


def nivel_convivencia(p: dict) -> int:
    npi     = p.get("npi")      or 0
    barthel = p.get("barthel")
    necpal  = p.get("necpal")   or 0
    norton  = p.get("norton")
    meclobo = p.get("meclobo")
    tinetti = p.get("tinetti")
    gds     = p.get("gds")      or 0
    goldberg= p.get("goldberg") or 0

    if npi >= 4:
        return 1

    if barthel is not None:
        if barthel <= 20:
            return 2
        if barthel <= 40 and necpal >= 1:
            return 2
        if barthel <= 40:
            return 3
        sev = (necpal >= 1
               or (norton  is not None and norton  <= 12)
               or (meclobo is not None and meclobo <= 10))
        if barthel <= 60 and sev:
            return 3
        if barthel <= 60:
            return 4

    if tinetti is not None and tinetti < 10:
        return 4

    if gds >= 5 or goldberg >= 4:
        return 5

    return 6


# MENÚ DIETAS

_DIET_MENU = {
    "Diabético":          "Menú Diabético",
    "Celíaco":            "Sin Gluten",
    "IntoleranteLactosa": "Sin Lactosa",
    "Hipocalórica":       "Hipocalórico",
    "BajoEnSal":          "Hiposódico",
    "NoMasticable":       "Triturado/Puré",
    "Vegetariano":        "Vegetariano",
    "Vegano":             "Vegano",
    "RicaEnFibra":        "Rico en Fibra",
    "DietaBlanda":        "Blanda",
}


# CLASE: TFG_CargadorOntologia

class TFG_CargadorOntologia:
    """Carga el grafo TTL y el suplemento XLS, y extrae los perfiles de residentes."""

    def _uri_propiedad(self, local: str) -> URIRef:
        return URIRef(H + local)

    def _leer_booleano(self, grafo, sujeto, propiedad):
        """Devuelve True/False si la propiedad existe, o None si no está en el grafo."""
        for valor in grafo.objects(sujeto, self._uri_propiedad(propiedad)):
            return str(valor).lower() == "true"
        return None

    def _leer_entero(self, grafo, sujeto, propiedad):
        """Devuelve el valor numérico de la propiedad, o None si no existe."""
        for valor in grafo.objects(sujeto, self._uri_propiedad(propiedad)):
            try:
                return int(valor)
            except:
                pass
        return None

    def _leer_texto(self, grafo, sujeto, propiedad):
        """Devuelve el texto de la propiedad, o None si no existe."""
        for valor in grafo.objects(sujeto, self._uri_propiedad(propiedad)):
            return str(valor)
        return None

    def _leer_objetos_relacionados(self, grafo, sujeto, propiedad) -> set[str]:
        """Devuelve un conjunto con los nombres de los objetos relacionados."""
        resultado = set()
        for valor in grafo.objects(sujeto, self._uri_propiedad(propiedad)):
            # Quedarse solo con la parte final del IRI (después del / o #)
            nombre = str(valor).rsplit("/", 1)[-1].rsplit("#", 1)[-1]
            resultado.add(nombre)
        return resultado

    def cargar_grafo_ttl(self, ruta: Path) -> Graph:
        if not ruta.exists():
            sys.exit(f"\nERROR: '{ruta}' no encontrado.\n"
                     f"Coloca el fichero Turtle en: {ruta.parent}\n")
        grafo = Graph()
        grafo.parse(str(ruta), format="turtle")
        print(f"  ✓ Grafo: {len(grafo):,} tripletas  ←  '{ruta.name}'")
        return grafo

    def cargar_datos_excel_xls(self) -> dict[str, dict]:
        """Lee el XLS y devuelve {codigo: {hobby, religion}} para los 2 campos nuevos."""
        if XLS_FILE is None:
            print("  ResidenciaDB-xlsx.xls no encontrado → hobby/religión no disponibles")
            return {}
        try:
            import xlrd
        except ImportError:
            print("   xlrd no instalado (pip install xlrd) → hobby/religión no disponibles")
            return {}

        wb  = xlrd.open_workbook(str(XLS_FILE))
        ws  = wb.sheet_by_index(0)
        hdr = [ws.cell_value(0, j) for j in range(ws.ncols)]

        def col(name):
            for j, h in enumerate(hdr):
                if h.lower().replace("ñ","n").replace("é","e") == name.lower().replace("ñ","n").replace("é","e"):
                    return j
            return None

        c_id      = col("ID")
        c_hobby   = col("tieneHobby")
        c_relig   = col("practicaReligion")

        datos_excel_xls: dict[str, dict] = {}
        for r in range(1, ws.nrows):
            code = str(ws.cell_value(r, c_id)).strip()
            hobby = str(ws.cell_value(r, c_hobby)).strip() if c_hobby is not None else ""
            relig = str(ws.cell_value(r, c_relig)).strip() if c_relig is not None else ""
            datos_excel_xls[code] = {"hobby": hobby, "religion": relig}

        print(f"  ✓ Excel suplementario: {len(datos_excel_xls)} registros  ←  '{XLS_FILE.name}'")
        return datos_excel_xls

    def _extraer_perfil_individual(self, grafo: Graph, codigo: str) -> dict:
        per = URIRef(D + "Persona/"   + codigo)
        res = URIRef(D + "Residente/" + codigo)
        return {
            "codigo"          : codigo,
            "nombre"          : self._leer_texto(grafo, per, "nombre") or codigo,
            "edad"            : self._leer_entero(grafo, per, "tieneEdad"),
            "genero"          : self._leer_texto(grafo, per, "género"),
            "barthel"         : self._leer_entero(grafo, res, "puntuacionBarthel"),
            "gds"             : self._leer_entero(grafo, res, "puntuacionGds"),
            "goldberg"        : self._leer_entero(grafo, res, "puntuacionGoldberg"),
            "meclobo"         : self._leer_entero(grafo, res, "puntuacionMecLobo"),
            "norton"          : self._leer_entero(grafo, res, "puntuacionNorton"),
            "tinetti"         : self._leer_entero(grafo, res, "puntuacionTinetti"),
            "pfeiffer"        : self._leer_entero(grafo, res, "puntuacionPfeiffer"),
            "npi"             : self._leer_entero(grafo, res, "puntuacionNpi"),
            "necpal"          : self._leer_entero(grafo, res, "puntuacionNecpal"),
            "colesterol"      : self._leer_entero(grafo, res, "nivelColesterol"),
            "presionSistolica": self._leer_entero(grafo, res, "presionSistolica"),
            "heartRate"       : self._leer_entero(grafo, res, "Heart_Rate"),
            "esFumador"       : self._leer_booleano(grafo, res, "esFumador"),
            "tieneDiabetes"   : self._leer_booleano(grafo, res, "tieneDiabetes"),
            "tieneObesidad"   : self._leer_booleano(grafo, res, "tieneObesidad"),
            "consumeAlcohol"  : self._leer_booleano(grafo, res, "consumeAlcohol"),
            "perdidaMemoria"  : self._leer_booleano(grafo, res, "tienePerdidaMemoria"),
            "usaSillaRuedas"  : self._leer_booleano(grafo, res, "usaSillaRuedas"),
            "usaPaniales"     : self._leer_booleano(grafo, res, "usaPañales"),
            "ayudaCaminar"    : self._leer_booleano(grafo, res, "ayudaCaminar"),
            "ayudaComer"      : self._leer_booleano(grafo, res, "ayudaComer"),
            "ayudaVestir"     : self._leer_booleano(grafo, res, "ayudaVestir"),
            "necesitaAsistNoc": self._leer_booleano(grafo, res, "necesitaAsistenciaNocturna"),
            "quiereHabComp"   : self._leer_booleano(grafo, res, "quiereHabitacionCompartida"),
            "recibeVisitas"   : self._leer_booleano(grafo, res, "recibeVisitasFrecuentes"),
            "tieneMascota"    : self._leer_booleano(grafo, res, "tieneMascota"),
            "toleraMascotas"  : self._leer_booleano(grafo, res, "toleraMascotas"),
            "viviaSolo"       : self._leer_booleano(grafo, res, "viviaSolo"),
            "frecVisitas"     : self._leer_texto(grafo, res, "frecuenciaVisitasFamiliares"),
            "diasActividad"   : self._leer_entero(grafo, res, "diasActividadFisica"),
            "nivelEstres"     : self._leer_entero(grafo, res, "nivelEstres"),
            "horasSueno"      : self._leer_entero(grafo, res, "horasDeSueno"),
            "presupuesto"     : self._leer_entero(grafo, res, "presupuesto"),
            "idiomas"         : self._leer_objetos_relacionados(grafo, res, "hablaIdioma"),
            "cicloSueno"      : self._leer_objetos_relacionados(grafo, res, "tieneCicloSueño"),
            "ruido"           : self._leer_objetos_relacionados(grafo, res, "toleraNivelRuido"),
            "ruidoTol"        : self._leer_objetos_relacionados(grafo, res, "nivelToleranciaRuido"),
            "zonasComunes"    : self._leer_objetos_relacionados(grafo, res, "frecuentaZonaComun"),
            "nivelActividad"  : self._leer_objetos_relacionados(grafo, res, "tieneNivelActividad"),
            "dieta"           : self._leer_objetos_relacionados(grafo, res, "tieneDietaEspecial"),
            "mascota"         : self._leer_objetos_relacionados(grafo, res, "leGustaMascota"),
            "necesidadMedica" : self._leer_objetos_relacionados(grafo, res, "tieneNecesidadMedica"),
            "riesgoSalud"     : self._leer_objetos_relacionados(grafo, res, "tieneRiesgoSalud"),
            "afinidadMedica"  : self._leer_objetos_relacionados(grafo, res, "tieneAfinidadMedicaCon"),
            "prefiereGenero"  : set(),
            "familiarDe"      : self._leer_objetos_relacionados(grafo, res, "esFamiliarDe"),
            "casadoCon"       : self._leer_objetos_relacionados(grafo, res, "estaCasadoCon"),
            "rechazaA"        : self._leer_objetos_relacionados(grafo, res, "rechazaConvivirCon"),
        }

    def cargar_todos_perfiles(self, grafo: Graph, datos_excel_xls: dict[str, dict]) -> dict[str, dict]:
        RES       = URIRef(H + "Residente")
        _prop_gen = URIRef(H + "prefiereCompañeroDeGénero")
        codigos: set[str] = set()
        for s in grafo.subjects(RDF.type, RES):
            iri = str(s)
            if iri.startswith(D):
                codigos.add(iri.rsplit("/", 1)[-1])

        perfiles: dict[str, dict] = {}
        for code in sorted(codigos):
            p = self._extraer_perfil_individual(grafo, code)
            res_uri = URIRef(D + "Residente/" + code)
            p["prefiereGenero"] = {
                str(o).rsplit("/",1)[-1].rsplit("#",1)[-1]
                for o in grafo.objects(res_uri, _prop_gen)
            }
            # Campos suplementarios del Excel
            xls = datos_excel_xls.get(code, {})
            p["hobby"]    = xls.get("hobby", "")
            p["religion"] = xls.get("religion", "")
            p["nivel"] = nivel_convivencia(p)
            perfiles[code] = p

        print(f"  ✓ {len(perfiles)} residentes cargados")
        return perfiles



# CLASE: TFG_MotorCompatibilidad

class TFG_MotorCompatibilidad:
    """Calcula puntuaciones de afinidad entre pares de residentes."""

    def _es_madrugador(self, ciclo: set) -> bool:
        """Comprueba si el ciclo de sueño del residente es de madrugador."""
        for valor in ciclo:
            if "Madrug" in valor:
                return True
        return False

    def _es_noctambulo(self, ciclo: set) -> bool:
        """Comprueba si el ciclo de sueño del residente es de noctámbulo."""
        for valor in ciclo:
            if "Noct" in valor:
                return True
        return False

    def _orden_nivel_ruido(self, niveles: set) -> int:
        """Devuelve el orden numérico del nivel de ruido (0=silencioso, 4=alto)."""
        for valor in niveles:
            if valor in ORDEN_RUIDO:
                return ORDEN_RUIDO[valor]
        return None

    def _preferencia_genero_compatible(self, preferencia: set, genero_compañero: str) -> bool:
        """Comprueba si el género del compañero encaja con la preferencia del residente."""
        # Si no hay preferencia declarada, es compatible con cualquiera
        if not preferencia or genero_compañero is None:
            return True
        for pref in preferencia:
            genero_esperado = GENERO_MAP.get(pref)
            if genero_esperado is None:
                return True  # preferencia desconocida  (no bloquear)
            if genero_compañero.lower() == genero_esperado.lower():
                return True
        return False

    def _verificar_incompatibilidades_veto(self, perfil_a: dict, perfil_b: dict) -> tuple[bool, str]:
        if perfil_b["codigo"] in perfil_a["rechazaA"] or perfil_a["codigo"] in perfil_b["rechazaA"]:
            return True, "Rechazo explícito declarado"
        if (perfil_a["esFumador"] is not None and perfil_b["esFumador"] is not None
                and perfil_a["esFumador"] != perfil_b["esFumador"]):
            return True, "Fumador ↔ No fumador"
        if perfil_a["tieneMascota"] and perfil_b["toleraMascotas"] is False:
            return True, f"{perfil_a['nombre']} tiene mascota / {perfil_b['nombre']} no tolera mascotas"
        if perfil_b["tieneMascota"] and perfil_a["toleraMascotas"] is False:
            return True, f"{perfil_b['nombre']} tiene mascota / {perfil_a['nombre']} no tolera mascotas"
        if perfil_a["recibeVisitas"] and "Silencioso" in perfil_b["ruido"] and perfil_b["quiereHabComp"]:
            return True, f"{perfil_a['nombre']} recibe visitas ↔ {perfil_b['nombre']} prefiere silencio"
        if perfil_b["recibeVisitas"] and "Silencioso" in perfil_a["ruido"] and perfil_a["quiereHabComp"]:
            return True, f"{perfil_b['nombre']} recibe visitas ↔ {perfil_a['nombre']} prefiere silencio"
        return False, ""

    def calcular_afinidad_par(self, perfil_a: dict, perfil_b: dict) -> tuple[float, list]:
        detalle: list[tuple[str, int]] = []
        score = 0

        vetado, motivo = self._verificar_incompatibilidades_veto(perfil_a, perfil_b)
        if vetado:
            return 0.0, [("VETO: " + motivo, 0)]

        def add(label: str, key: str) -> None:
            w = PESO[key]; detalle.append((label, w))
            nonlocal score; score += w

        if perfil_b["codigo"] in perfil_a["casadoCon"] or perfil_a["codigo"] in perfil_b["casadoCon"]:
            add("Pareja (casados)", "casado_con")
        elif perfil_b["codigo"] in perfil_a["familiarDe"] or perfil_a["codigo"] in perfil_b["familiarDe"]:
            add("Familiar", "es_familiar")

        if perfil_a["idiomas"] & perfil_b["idiomas"]:
            add("Mismo idioma", "mismo_idioma")

        if perfil_a["cicloSueno"] and perfil_b["cicloSueno"]:
            am, an = self._es_madrugador(perfil_a["cicloSueno"]), self._es_noctambulo(perfil_a["cicloSueno"])
            bm, bn = self._es_madrugador(perfil_b["cicloSueno"]), self._es_noctambulo(perfil_b["cicloSueno"])
            if (am and bm) or (an and bn):
                add("Mismo ciclo sueño", "mismo_ciclo_sueno")
            elif (am and bn) or (an and bm):
                add("Ciclo sueño opuesto", "ciclo_opuesto")

        if perfil_a["horasSueno"] is not None and perfil_b["horasSueno"] is not None:
            if abs(perfil_a["horasSueno"] - perfil_b["horasSueno"]) <= 1:
                add("Horas sueño similares", "horas_sueno_similares")

        if perfil_a["ruido"] and perfil_b["ruido"]:
            if perfil_a["ruido"] & perfil_b["ruido"]:
                add("Mismo nivel de ruido tolerado", "mismo_nivel_ruido")
            else:
                ra, rb = self._orden_nivel_ruido(perfil_a["ruido"]), self._orden_nivel_ruido(perfil_b["ruido"])
                if ra is not None and rb is not None and abs(ra - rb) >= 3:
                    add("Ruido muy incompatible", "ruido_incompatible")
        if perfil_a["ruidoTol"] and perfil_b["ruidoTol"]:
            ra, rb = self._orden_nivel_ruido(perfil_a["ruidoTol"]), self._orden_nivel_ruido(perfil_b["ruidoTol"])
            if ra is not None and rb is not None and abs(ra - rb) >= 2:
                add("Tolerancia al ruido diferente", "ruido_tol_diferente")

        na, nb = nivel_barthel(perfil_a["barthel"]), nivel_barthel(perfil_b["barthel"])
        if na is not None and nb is not None:
            diff_dep = abs(na - nb)
            if diff_dep == 0:
                add(f"Mismo nivel dependencia ({ETIQ_BARTHEL.get(na,'')})", "dependencia_igual")
            elif diff_dep == 1:
                add("Nivel dependencia similar", "dependencia_1nivel")
            elif diff_dep == 2:
                add("Niveles dependencia distintos", "dependencia_2nivel")
            else:
                add("Niveles dependencia muy distintos", "dependencia_3nivel")

        nmc_a, nmc_b = nivel_meclobo(perfil_a["meclobo"]), nivel_meclobo(perfil_b["meclobo"])
        if nmc_a is not None and nmc_b is not None:
            diff_cog = abs(nmc_a - nmc_b)
            if diff_cog == 0:
                add(f"Mismo nivel cognitivo ({ETIQ_MECLOBO.get(nmc_a,'')})", "cognitivo_igual")
            elif diff_cog >= 2:
                add("Estado cognitivo muy diferente", "cognitivo_2nivel")

        rt_a, rt_b = riesgo_tinetti(perfil_a["tinetti"]), riesgo_tinetti(perfil_b["tinetti"])
        if rt_a == rt_b == "Bajo":
            add("Ambos con bajo riesgo de caída", "ambos_bajo_riesgo_caida")
        elif rt_a == rt_b == "Alto":
            add("Ambos con alto riesgo de caída", "ambos_alto_riesgo_caida")

        adl_a, adl_b = n_adl(perfil_a), n_adl(perfil_b)
        diff_adl = abs(adl_a - adl_b)
        if diff_adl == 0:
            add(f"Mismo nivel de ayudas AVD ({adl_a}/3)", "adl_igual")
        elif diff_adl >= 2:
            add("Ayudas AVD muy diferentes", "adl_muy_diferente")

        if perfil_a["necesitaAsistNoc"] is not None and perfil_b["necesitaAsistNoc"] is not None:
            if perfil_a["necesitaAsistNoc"] and perfil_b["necesitaAsistNoc"]:
                add("Ambos necesitan asistencia nocturna", "ambos_asistencia_noc")
            elif perfil_a["necesitaAsistNoc"] != perfil_b["necesitaAsistNoc"]:
                add("Asistencia nocturna discordante", "asistencia_noc_dif")

        if perfil_a["tieneDiabetes"] and perfil_b["tieneDiabetes"]:
            add("Ambos diabéticos", "ambos_diabeticos")
        if perfil_a["consumeAlcohol"] is False and perfil_b["consumeAlcohol"] is False:
            add("Ambos no consumen alcohol", "ambos_no_alcohol")
        if perfil_a["perdidaMemoria"] and perfil_b["perdidaMemoria"]:
            add("Ambos con pérdida de memoria", "perdida_memoria_doble")

        shared_nm = (perfil_a["necesidadMedica"] - {"Ninguna"}) & (perfil_b["necesidadMedica"] - {"Ninguna"})
        for nm in list(shared_nm)[:3]:
            w = PESO["necesidad_medica_comun"]
            detalle.append((f"Necesidad médica común: {nm}", w)); score += w

        shared_am = (perfil_a["afinidadMedica"] - {"Ninguna"}) & (perfil_b["afinidadMedica"] - {"Ninguna"})
        for am in list(shared_am)[:2]:
            w = PESO["afinidad_medica_comun"]
            detalle.append((f"Protocolo médico común: {am}", w)); score += w

        if perfil_a["nivelActividad"] & perfil_b["nivelActividad"]:
            add("Mismo nivel de actividad", "mismo_nivel_actividad")
        if perfil_a["diasActividad"] is not None and perfil_b["diasActividad"] is not None:
            diff_act = abs(perfil_a["diasActividad"] - perfil_b["diasActividad"])
            if diff_act <= 1:
                add("Actividad física similar", "actividad_fisica_sim")
            elif diff_act > 4:
                add("Actividad física muy diferente", "actividad_muy_diferente")

        if perfil_a["zonasComunes"] & perfil_b["zonasComunes"]:
            add("Misma zona común frecuentada", "misma_zona_comun")

        if perfil_a["quiereHabComp"] and perfil_b["quiereHabComp"]:
            add("Ambos quieren habitación compartida", "ambos_hab_compartida")
        elif (perfil_a["quiereHabComp"] is not None and perfil_b["quiereHabComp"] is not None
              and perfil_a["quiereHabComp"] != perfil_b["quiereHabComp"]):
            add("Discrepancia en habitación compartida", "hab_discordante")

        if (self._preferencia_genero_compatible(perfil_a["prefiereGenero"], perfil_b["genero"]) and
                self._preferencia_genero_compatible(perfil_b["prefiereGenero"], perfil_a["genero"])):
            add("Preferencia de género compatible", "genero_compatible")

        if perfil_a["frecVisitas"] and perfil_b["frecVisitas"] and perfil_a["frecVisitas"] == perfil_b["frecVisitas"]:
            add(f"Misma frecuencia de visitas ({perfil_a['frecVisitas']})", "misma_frec_visitas")

        if perfil_a["viviaSolo"] is not None and perfil_b["viviaSolo"] is not None:
            if perfil_a["viviaSolo"] == perfil_b["viviaSolo"]:
                add("Ambos vivían solos" if perfil_a["viviaSolo"] else "Ambos vivían acompañados",
                    "mismo_vivio_solo")

        d_a = perfil_a["dieta"] - {"Ninguna"}; d_b = perfil_b["dieta"] - {"Ninguna"}
        if d_a and d_b and (d_a & d_b):
            add(f"Misma dieta especial ({', '.join(d_a & d_b)})", "misma_dieta")

        m_a = perfil_a["mascota"] - {"Ninguna"}; m_b = perfil_b["mascota"] - {"Ninguna"}
        if m_a and m_b and (m_a & m_b):
            add(f"Misma afinidad mascotas ({', '.join(m_a & m_b)})", "mascota_afinidad")

        if perfil_a["nivelEstres"] is not None and perfil_b["nivelEstres"] is not None:
            diff_e = abs(perfil_a["nivelEstres"] - perfil_b["nivelEstres"])
            if diff_e <= 2:
                add("Nivel de estrés similar", "estres_similar")
            elif diff_e > 5:
                add("Nivel de estrés muy diferente", "estres_muy_diferente")

        if perfil_a["presupuesto"] is not None and perfil_b["presupuesto"] is not None:
            diff_p = abs(perfil_a["presupuesto"] - perfil_b["presupuesto"])
            if diff_p <= 500:
                add("Presupuesto similar (±500 €)", "presupuesto_similar")
            elif diff_p > 1500:
                add("Presupuesto muy diferente (>1500 €)", "presupuesto_muy_dif")

        if perfil_a["edad"] is not None and perfil_b["edad"] is not None:
            if abs(perfil_a["edad"] - perfil_b["edad"]) > 20:
                add("Diferencia de edad > 20 años", "edad_muy_diferente")

        if perfil_a["goldberg"] is not None and perfil_b["goldberg"] is not None:
            if perfil_a["goldberg"] >= 4 and perfil_b["goldberg"] >= 4:
                add("Ambos con ansiedad elevada (Goldberg ≥ 4)", "ansiedad_doble")
        if perfil_a["gds"] is not None and perfil_b["gds"] is not None:
            if perfil_a["gds"] >= 5 and perfil_b["gds"] >= 5:
                add("Ambos con depresión elevada (GDS ≥ 5)", "depresion_doble")

        # Hobby (tieneHobby) 
        if perfil_a["hobby"] and perfil_b["hobby"] and perfil_a["hobby"] == perfil_b["hobby"]:
            add(f"Misma afición: {perfil_a['hobby']}", "mismo_hobby")

        # Religión (practicaReligion)
        if perfil_a["religion"] and perfil_b["religion"]:
            if perfil_a["religion"] == perfil_b["religion"]:
                add(f"Misma práctica religiosa: {perfil_a['religion']}", "misma_religion")
            else:
                
                opuestas = {"Ateísmo", "Catolicismo"}
                if {perfil_a["religion"], perfil_b["religion"]} == opuestas:
                    add("Práctica religiosa opuesta (Ateísmo ↔ Catolicismo)", "religion_opuesta")

        score = max(0, min(100, score))
        return round(score, 1), detalle

    def calcular_matriz_afinidad_completa(self, perfiles: dict) -> tuple:
        """Calcula la afinidad entre todos los pares posibles de residentes."""
        codigos = sorted(perfiles)

        # Guardar la puntuación y el detalle de cada par
        scores = {}
        detalles = {}

        for i, codigo_a in enumerate(codigos):
            for codigo_b in codigos[i+1:]:
                puntuacion, detalle = self.calcular_afinidad_par(perfiles[codigo_a], perfiles[codigo_b])
                # Guardar en ambas direcciones 
                scores[(codigo_a, codigo_b)] = puntuacion
                scores[(codigo_b, codigo_a)] = puntuacion
                detalles[(codigo_a, codigo_b)] = detalle
                detalles[(codigo_b, codigo_a)] = detalle

        # Construir la matriz de afinidad como tabla 
        filas_matriz = []
        for codigo_a in codigos:
            fila = []
            for codigo_b in codigos:
                if codigo_a == codigo_b:
                    fila.append(100.0)  
                else:
                    fila.append(scores.get((codigo_a, codigo_b), 0.0))
            filas_matriz.append(fila)
        matriz_afinidad = pd.DataFrame(filas_matriz, index=codigos, columns=codigos)

        # Lista de parejas ordenadas de mayor a menor afinidad
        parejas = []
        for i, codigo_a in enumerate(codigos):
            for codigo_b in codigos[i+1:]:
                parejas.append({
                    "A":        codigo_a,
                    "nombre_A": perfiles[codigo_a]["nombre"],
                    "nivel_A":  perfiles[codigo_a]["nivel"],
                    "B":        codigo_b,
                    "nombre_B": perfiles[codigo_b]["nombre"],
                    "nivel_B":  perfiles[codigo_b]["nivel"],
                    "pct":      scores[(codigo_a, codigo_b)],
                    "detalle":  detalles[(codigo_a, codigo_b)],
                })
        parejas.sort(key=lambda x: x["pct"], reverse=True)

        # Ranking individual: para cada residente, sus candidatos ordenados
        ranking = {}
        for codigo_a in codigos:
            candidatos = []
            for codigo_b in codigos:
                if codigo_b != codigo_a:
                    candidatos.append((
                        scores.get((codigo_a, codigo_b), 0.0),
                        codigo_b,
                        detalles.get((codigo_a, codigo_b), [])
                    ))
            candidatos.sort(key=lambda x: x[0], reverse=True)
            ranking[codigo_a] = candidatos

        return matriz_afinidad, parejas, ranking



# CLASE: TFG_ExportadorResultados

class TFG_ExportadorResultados:
    """Genera todos los archivos de salida (Excel, CSV, TTL, alertas, log XAI)."""

    def exportar_excel_profesional(self, perfiles, matriz_afinidad, parejas, ranking, directorio):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                         GradientFill)
            from openpyxl.utils import get_column_letter
            from openpyxl.styles.numbers import FORMAT_PERCENTAGE
        except ImportError:
            print("  ⚠  openpyxl no instalado → saltando Excel (pip install openpyxl)")
            return

        directorio.mkdir(parents=True, exist_ok=True)
        codigos = sorted(perfiles)

        def solid(hex6): return PatternFill("solid", fgColor=hex6)

        def font(bold=False, color="000000", size=10, italic=False):
            return Font(bold=bold, color=color, size=size, italic=italic,
                        name="Calibri")

        def align(h="center", v="center", wrap=False):
            return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

        def border(color="CCCCCC"):
            s = Side(style="thin", color=color)
            return Border(left=s, right=s, top=s, bottom=s)

        NAVY       = solid("1A3A5C")
        NAVY2      = solid("2D6A9F")
        GREY_HDR   = solid("F2F2F2")
        WHITE      = solid("FFFFFF")
        ROW_ALT    = solid("EEF4FB")
        BRD        = border()
        BRD_HDR    = border("888888")

        def nivel_fill(niv): return solid(NIVEL_COLOR_BG[niv])
        def nivel_font(niv): return font(bold=True, color=NIVEL_COLOR_TXT[niv], size=9)

        def compat_fill(v):
            if v >= 100: return solid("D0D0D0")
            if v == 0:   return solid("FFCCCC")
            if v < 20:   return solid("FFE0B2")
            if v < 40:   return solid("FFF9C4")
            if v < 60:   return solid("C8E6C9")
            if v < 80:   return solid("66BB6A")
            return            solid("2E7D32")

        def compat_font(v):
            return font(bold=(v >= 60 and v < 100),
                        color="FFFFFF" if v >= 80 else "000000",
                        size=9)

        def bool_str(v):
            if v is True:  return "Si"
            if v is False: return "No"
            return "—"

        wb = Workbook()

       
        # HOJA 1  (FICHAS CLÍNICAS DE RESIDENTES)
      
        ws1 = wb.active; ws1.title = "Fichas Residentes"
        ws1.sheet_view.showGridLines = False

        # Título principal
        ws1.merge_cells("A1:AI1")
        tit = ws1["A1"]
        tit.value = "SISTEMA CO-HOUSING SENIOR  ·  Fichas Clínicas de Residentes  ·  TFG · Álvaro · 2026"
        tit.fill = NAVY
        tit.font = font(bold=True, color="FFFFFF", size=13)
        tit.alignment = align()
        ws1.row_dimensions[1].height = 28

        # Subtitulos
        secciones = [
            ("A2:F2", "DATOS PERSONALES",    "1A3A5C"),
            ("G2:G2", "NIVEL",               "2D6A9F"),
            ("H2:V2", "ESCALAS CLÍNICAS (puntuación → interpretación)", "16537E"),
            ("W2:AC2","CONDICIONES MÉDICAS", "1A3A5C"),
            ("AD2:AE2","AFICIONES / VALORES", "2D6A9F"),
            ("AF2:AK2","ESTILO DE VIDA",     "1A3A5C"),
        ]
        for rng, txt, col in secciones:
            ws1.merge_cells(rng)
            c = ws1[rng.split(":")[0]]
            c.value = txt; c.fill = solid(col)
            c.font = font(bold=True, color="FFFFFF", size=9)
            c.alignment = align()
        ws1.row_dimensions[2].height = 18

        #Fila de encabezados 
        HEADERS = [
            "#", "Código", "Nombre", "Edad", "Género", "Fecha ingreso",
            # Nivel
            "NIVEL DE CONVIVENCIA",
            # Escalas clínicas (9 escalas × 2 columnas = 18 columnas)
            "Barthel", "Dependencia (Barthel)",
            "MEC-Lobo", "Estado Cognitivo",
            "Tinetti", "Riesgo Caída",
            "GDS", "Depresión (GDS)",
            "Goldberg", "Ansiedad (Goldberg)",
            "Norton", "Riesgo Úlceras",
            "NPI", "Alteraciones Conducta",
            "Pfeiffer", "Función Cognitiva (Pfeiffer)",
            "NECPAL", "Pronóstico Vital",
            # Condiciones médicas (7 columnas)
            "AVD (0-3)", "Asist. Nocturna", "Silla Ruedas", "Diabetes",
            "Fumador", "Obesidad", "Pérd. Memoria",
            # Aficiones / Valores (2 columnas)
            "Hobby", "Religión",
            # Estilo de vida (6 columnas)
            "Ciclo Sueño", "Idioma(s)", "Nivel Actividad",
            "Presupuesto €", "Zonas Comunes", "Necesidades Médicas",
        ]
        for j, h in enumerate(HEADERS, 1):
            c = ws1.cell(3, j, h)
            c.fill = GREY_HDR
            c.font = font(bold=True, size=9)
            c.alignment = align(wrap=True)
            c.border = BRD_HDR
        ws1.row_dimensions[3].height = 32
        ws1.freeze_panes = "A4"

        # Anchos de columna
        col_widths = [5, 11, 14, 6, 8, 16,   
                      28,                       
                      9, 18, 9, 15, 9, 14,     
                      7, 15, 9, 15,            
                      9, 14, 7, 20, 7, 18,     
                      9, 14, 13, 10,           
                      10, 10, 13,              
                      16, 14,                  
                      14, 16, 16,             
                      12, 22, 30,             
                      ]
        for j, w in enumerate(col_widths, 1):
            ws1.column_dimensions[get_column_letter(j)].width = w

        # Filas de datos 
        for i, ca in enumerate(codigos):
            row = i + 4
            p   = perfiles[ca]
            niv = p["nivel"]
            nb  = nivel_barthel(p.get("barthel"))
            nc  = nivel_meclobo(p.get("meclobo"))
            alt = (i % 2 == 0)

            def cell(col, val, h="center", wrap=False, bold=False, italic=False,
                     fill=None, fnt=None, brd=BRD):
                c = ws1.cell(row, col, val)
                c.alignment = align(h=h, wrap=wrap)
                c.font = fnt or font(bold=bold, italic=italic, size=9)
                c.fill = fill or (ROW_ALT if alt else WHITE)
                if brd: c.border = brd
                return c

            cell(1, i+1)
            cell(2, ca, bold=True)
            cell(3, p["nombre"], h="left", bold=True)
            cell(4, p.get("edad") or "—")
            cell(5, p.get("genero") or "—")
            cell(6, "—")  # fecha ingreso (no está en ontología como campo directo)

            nv_c = ws1.cell(row, 7, NIVEL_LABEL[niv])
            nv_c.fill  = nivel_fill(niv)
            nv_c.font  = nivel_font(niv)
            nv_c.alignment = align(wrap=True)
            nv_c.border = border("666666")
            ws1.row_dimensions[row].height = 20

            # Escalas clínicas
            def scale_pair(col_num, raw_val, interp_val, pct=None):
                # puntuación
                sc = ws1.cell(row, col_num, raw_val if raw_val is not None else "—")
                sc.alignment = align()
                sc.font = font(bold=True, size=9)
                sc.fill = ROW_ALT if alt else WHITE
                sc.border = BRD
               
                ic = ws1.cell(row, col_num+1, interp_val)
                ic.alignment = align(h="left")
                ic.font = font(size=9, italic=True)
                ic.fill = ROW_ALT if alt else WHITE
                ic.border = BRD

            scale_pair(8,  p.get("barthel"),  ETIQ_BARTHEL.get(nb, "—") if nb is not None else "—")
            scale_pair(10, p.get("meclobo"),  ETIQ_MECLOBO.get(nc, "—") if nc is not None else "—")
            scale_pair(12, p.get("tinetti"),  riesgo_tinetti(p.get("tinetti")) or "—")
            scale_pair(14, p.get("gds"),      interp_gds(p.get("gds")))
            scale_pair(16, p.get("goldberg"), interp_goldberg(p.get("goldberg")))
            scale_pair(18, p.get("norton"),   interp_norton(p.get("norton")))
            scale_pair(20, p.get("npi"),      interp_npi(p.get("npi")))
            scale_pair(22, p.get("pfeiffer"), interp_pfeiffer(p.get("pfeiffer")))
            scale_pair(24, p.get("necpal"),   interp_necpal(p.get("necpal")))

            # Condiciones médicas
            for idx, (key, col) in enumerate([
                ("avd",             26),
                ("necesitaAsistNoc",27),
                ("usaSillaRuedas",  28),
                ("tieneDiabetes",   29),
                ("esFumador",       30),
                ("tieneObesidad",   31),
                ("perdidaMemoria",  32),
            ]):
                if key == "avd":
                    val = f"{n_adl(p)}/3"
                else:
                    val = bool_str(p.get(key))
                c = ws1.cell(row, col, val)
                c.alignment = align()
                c.font = font(size=9,
                              bold=(val == "Si"),
                              color=("C62828" if val == "Si" and key in
                                     ("esFumador","usaSillaRuedas","perdidaMemoria") else "000000"))
                c.fill = ROW_ALT if alt else WHITE
                c.border = BRD

            # Aficiones / Valores
            cell(33, p.get("hobby", "—") or "—", h="left")
            cell(34, p.get("religion", "—") or "—", h="left")
            # Estilo de vida
            cell(35, list(p["cicloSueno"])[0] if p["cicloSueno"] else "—", h="left")
            cell(36, ", ".join(p["idiomas"]) or "—", h="left")
            cell(37, ", ".join(p["nivelActividad"]) or "—", h="left")
            cell(38, f"{p['presupuesto']} €" if p.get("presupuesto") else "—")
            cell(39, ", ".join(p["zonasComunes"]) or "—", h="left", wrap=True)
            nm = ", ".join(p["necesidadMedica"] - {"Ninguna"}) or "Ninguna"
            cell(40, nm, h="left", wrap=True)

        # Bordes del bloque de datos 
        from openpyxl.styles import Border as OBorder, Side as OSide
        med = OSide(style="medium", color="1A3A5C")
        med_brd = OBorder(left=med, right=med, top=med, bottom=med)

        
        # HOJA 2  (MATRIZ DE AFINIDAD)
        
        ws2 = wb.create_sheet("Matriz Afinidad")
        ws2.sheet_view.showGridLines = False

        ws2.merge_cells(f"A1:{get_column_letter(len(codigos)+1)}1")
        t2 = ws2["A1"]
        t2.value = "MATRIZ DE AFINIDAD (%)  ·  Verde oscuro ≥ 80%  ·  Verde 60-79%  ·  Amarillo 40-59%  ·  Naranja < 40%  ·  Rojo = VETADO"
        t2.fill  = NAVY; t2.font = font(bold=True, color="FFFFFF", size=11)
        t2.alignment = align(); ws2.row_dimensions[1].height = 22

        # Cabecera columnas: código + nombre + nivel badge
        ws2.cell(2, 1, "↓ / →").fill = GREY_HDR
        ws2.cell(2, 1).font = font(bold=True, size=8)
        ws2.cell(2, 1).alignment = align()
        ws2.column_dimensions["A"].width = 16
        ws2.row_dimensions[2].height = 46

        for j, cb in enumerate(codigos, 2):
            pb  = perfiles[cb]
            niv = pb["nivel"]
            c   = ws2.cell(2, j,
                           f"{cb}\n{pb['nombre']}\n{NIVEL_SHORT[niv]}")
            c.fill      = nivel_fill(niv)
            c.font      = nivel_font(niv)
            c.alignment = align(wrap=True)
            c.border    = border("888888")
            ws2.column_dimensions[get_column_letter(j)].width = 10

        # Filas de datos
        for i, ca in enumerate(codigos, 3):
            pa  = perfiles[ca]
            niv = pa["nivel"]
            r   = ws2.cell(i, 1,
                           f"{ca}\n{pa['nombre']}\n{NIVEL_SHORT[niv]}")
            r.fill      = nivel_fill(niv)
            r.font      = nivel_font(niv)
            r.alignment = align(wrap=True)
            r.border    = border("888888")
            ws2.row_dimensions[i].height = 40

            for j, cb in enumerate(codigos, 2):
                v = float(matriz_afinidad.at[ca, cb])
                c = ws2.cell(i, j, "" if ca == cb else int(v))
                c.fill      = compat_fill(v)
                c.font      = compat_font(v)
                c.alignment = align()
                c.border    = border("AAAAAA")

        ws2.freeze_panes = "B3"

        
        # HOJA 3  (RANKING POR RESIDENTE)
        
        ws3 = wb.create_sheet("Ranking por Residente")
        ws3.sheet_view.showGridLines = False

        ws3.merge_cells("A1:U1")
        t3 = ws3["A1"]
        t3.value = "RANKING DE COMPATIBILIDAD POR RESIDENTE  ·  Todos los campos de la ontología"
        t3.fill = NAVY; t3.font = font(bold=True, color="FFFFFF", size=12)
        t3.alignment = align(); ws3.row_dimensions[1].height = 24

        H3 = ["#", "Código", "Nombre", "Edad", "Género", "NIVEL CONVIVENCIA",
              "Barthel / Dep.", "MEC-Lobo / Cogn.", "Tinetti / Caída",
              "GDS / Dep.", "Goldberg / Ans.", "Norton / Úlceras",
              "NPI / Conducta", "AVD", "N.Vetados",
              "Hobby", "Religión",
              "1º Match", "% 1º", "2º Match", "% 2º", "3º Match", "% 3º"]
        for j, h in enumerate(H3, 1):
            c = ws3.cell(2, j, h)
            c.fill = NAVY2; c.font = font(bold=True, color="FFFFFF", size=9)
            c.alignment = align(wrap=True); c.border = BRD_HDR
        ws3.row_dimensions[2].height = 36
        ws3.freeze_panes = "A3"

        w3_widths = [5, 11, 14, 6, 8, 28,
                     17, 17, 15, 14, 17, 16, 15, 7, 9,
                     16, 14,
                     18, 8, 18, 8, 18, 8]
        for j, w in enumerate(w3_widths, 1):
            ws3.column_dimensions[get_column_letter(j)].width = w

        for i, ca in enumerate(codigos, 3):
            p     = perfiles[ca]
            niv   = p["nivel"]
            nb    = nivel_barthel(p.get("barthel"))
            nc    = nivel_meclobo(p.get("meclobo"))
            fila_r = ranking.get(ca, [])
            n_vet  = sum(1 for s, _, _ in fila_r if s == 0)
            top3   = [(sc, perfiles[cb]["nombre"] + f"\n[{cb}] {NIVEL_SHORT[perfiles[cb]['nivel']]}",
                       sc) for sc, cb, _ in fila_r[:3] if sc > 0]
            alt    = (i % 2 == 0)
            bg     = ROW_ALT if alt else WHITE

            def w3cell(col, val, h="center", wrap=False, fnt=None, fill=None):
                c = ws3.cell(i, col, val)
                c.alignment = align(h=h, wrap=wrap)
                c.font  = fnt or font(size=9)
                c.fill  = fill or bg
                c.border = BRD
                return c

            w3cell(1,  i-2, fnt=font(bold=True, size=9))
            w3cell(2,  ca,  fnt=font(bold=True, size=9))
            w3cell(3,  p["nombre"], h="left")
            w3cell(4,  p.get("edad") or "—")
            w3cell(5,  p.get("genero") or "—")

            nc_cell = ws3.cell(i, 6, NIVEL_LABEL[niv])
            nc_cell.fill = nivel_fill(niv); nc_cell.font = nivel_font(niv)
            nc_cell.alignment = align(wrap=True); nc_cell.border = border("666666")

            w3cell(7,  f"{p.get('barthel','—')} → {ETIQ_BARTHEL.get(nb,'—') if nb is not None else '—'}", wrap=True)
            w3cell(8,  f"{p.get('meclobo','—')} → {ETIQ_MECLOBO.get(nc,'—') if nc is not None else '—'}", wrap=True)
            w3cell(9,  f"{p.get('tinetti','—')} → {riesgo_tinetti(p.get('tinetti')) or '—'}", wrap=True)
            w3cell(10, f"{p.get('gds','—')} ({interp_gds(p.get('gds'))})", wrap=True)
            w3cell(11, f"{p.get('goldberg','—')} ({interp_goldberg(p.get('goldberg'))})", wrap=True)
            w3cell(12, f"{p.get('norton','—')} ({interp_norton(p.get('norton'))})", wrap=True)
            w3cell(13, f"{p.get('npi','—')} ({interp_npi(p.get('npi'))})", wrap=True)
            w3cell(14, f"{n_adl(p)}/3")
            w3cell(15, n_vet, fnt=font(bold=True, color=("C62828" if n_vet > 30 else "000000"), size=9))
            w3cell(16, p.get("hobby", "—") or "—", h="left")
            w3cell(17, p.get("religion", "—") or "—", h="left")

            for k, (sc, lbl, _) in enumerate(top3):
                col_lbl = 18 + k*2
                col_pct = 19 + k*2
                ml = ws3.cell(i, col_lbl, lbl)
                ml.alignment = align(wrap=True); ml.font = font(size=8)
                ml.fill = bg; ml.border = BRD
                pc = ws3.cell(i, col_pct, sc)
                pc.fill = compat_fill(sc); pc.font = compat_font(sc)
                pc.alignment = align(); pc.border = BRD

            ws3.row_dimensions[i].height = 32

        # HOJA 4 (TOP 50 PAREJAS GLOBALES)
    
        ws4 = wb.create_sheet("Top Parejas Global")
        ws4.sheet_view.showGridLines = False

        ws4.merge_cells("A1:I1")
        t4 = ws4["A1"]
        t4.value = "TOP 50 PAREJAS MÁS COMPATIBLES  ·  Ranking global por porcentaje de afinidad"
        t4.fill = NAVY; t4.font = font(bold=True, color="FFFFFF", size=12)
        t4.alignment = align(); ws4.row_dimensions[1].height = 24

        H4 = ["#", "Residente A", "Nombre A", "Nivel A",
              "Residente B", "Nombre B", "Nivel B",
              "Afinidad %", "Razones de compatibilidad"]
        for j, h in enumerate(H4, 1):
            c = ws4.cell(2, j, h)
            c.fill = NAVY2; c.font = font(bold=True, color="FFFFFF", size=10)
            c.alignment = align(wrap=True); c.border = BRD_HDR
        ws4.row_dimensions[2].height = 28
        ws4.freeze_panes = "A3"

        w4_widths = [5, 12, 14, 30, 12, 14, 30, 12, 65]
        for j, w in enumerate(w4_widths, 1):
            ws4.column_dimensions[get_column_letter(j)].width = w

        top50 = [p for p in parejas if p["pct"] > 0][:50]
        for i, par in enumerate(top50, 3):
            alt = (i % 2 == 0)
            bg  = ROW_ALT if alt else WHITE
            nA, nB = par["nivel_A"], par["nivel_B"]
            razones = " · ".join(
                txt for txt, pts in
                sorted(par["detalle"], key=lambda x: -x[1])[:5]
                if pts > 0
            )

            def w4cell(col, val, h="center", wrap=False, fnt=None, fill=None):
                c = ws4.cell(i, col, val)
                c.alignment = align(h=h, wrap=wrap)
                c.font  = fnt or font(size=10)
                c.fill  = fill or bg
                c.border = BRD
                return c

            w4cell(1, i-2, fnt=font(bold=True, size=10))
            w4cell(2, par["A"], fnt=font(bold=True, size=10))
            w4cell(3, par["nombre_A"], h="left")
            cA = ws4.cell(i, 4, NIVEL_LABEL[nA])
            cA.fill = nivel_fill(nA); cA.font = nivel_font(nA)
            cA.alignment = align(wrap=True); cA.border = BRD

            w4cell(5, par["B"], fnt=font(bold=True, size=10))
            w4cell(6, par["nombre_B"], h="left")
            cB = ws4.cell(i, 7, NIVEL_LABEL[nB])
            cB.fill = nivel_fill(nB); cB.font = nivel_font(nB)
            cB.alignment = align(wrap=True); cB.border = BRD

            pct_c = ws4.cell(i, 8, par["pct"])
            pct_c.fill = compat_fill(par["pct"]); pct_c.font = compat_font(par["pct"])
            pct_c.alignment = align(); pct_c.border = border("888888")
            ws4.cell(i, 8).font = Font(bold=True, size=11, name="Calibri",
                                       color=("FFFFFF" if par["pct"] >= 80 else "000000"))

            w4cell(9, razones, h="left", wrap=True, fnt=font(size=9, italic=True))
            ws4.row_dimensions[i].height = 28

        # Guardar 
        ruta_x = directorio / "fichas_residentes.xlsx"
        try:
            wb.save(str(ruta_x))
            print(f"  ✓ {ruta_x.name}  (4 hojas: Fichas · Matriz · Ranking · Top50)")
        except PermissionError:
            import time
            nombre_alt = directorio / f"fichas_residentes_{int(time.time())}.xlsx"
            wb.save(str(nombre_alt))
            print(f"   fichas_residentes.xlsx estaba abierto en Excel.")
            print(f"    Guardado como: {nombre_alt.name}")


    def exportar_grafo_semantico_enriquecido(self, grafo_original: Graph, perfiles: dict, matriz_afinidad, directorio: Path) -> None:
        """Serializa el grafo original + todas las tripletas inferidas en Turtle."""
        from rdflib import Graph as RG, URIRef as URI, Literal, Namespace, RDF as _RDF
        from rdflib.namespace import OWL, RDFS, XSD

        directorio.mkdir(parents=True, exist_ok=True)
        gi = RG()

        # Copiar grafo original
        for triple in grafo_original:
            gi.add(triple)

        Hns = Namespace(H)
        Dns = Namespace(D)
        gi.bind("h", Hns)
        gi.bind("d", Dns)
        gi.bind("owl", OWL)
        gi.bind("rdfs", RDFS)
        gi.bind("xsd", XSD)

        _RES    = URI(H + "Residente")
        _NIVEL  = URI(H + "tieneNivelConvivencia")
        _COMPAT = URI(H + "esCompatibleCon")
        _INCOMPAT  = URI(H + "esIncompatibleCon")
        _DIETA  = URI(H + "comparteDietaCon")
        _AFICION = URI(H + "comparteAficionCon")
        _HOBBY  = URI(H + "tieneHobby")
        _RELIG  = URI(H + "practicaReligion")

        nivel_iris = {n: URI(H + f"Nivel{n}") for n in range(1, 7)}
        codigos = sorted(perfiles)

        n_nivel = n_compat = n_incompat = n_dieta = n_aficion = n_hobby = n_relig = 0

        for code, p in perfiles.items():
            res_uri = URI(D + "Residente/" + code)

            # tieneNivelConvivencia
            gi.add((res_uri, _NIVEL, nivel_iris[p["nivel"]]))
            n_nivel += 1

            # tieneHobby 
            if p.get("hobby"):
                gi.add((res_uri, _HOBBY, Literal(p["hobby"])))
                n_hobby += 1

            # practicaReligion 
            if p.get("religion"):
                gi.add((res_uri, _RELIG, Literal(p["religion"])))
                n_relig += 1

        # Pares de compatibilidad e incompatibilidad
        for i, ca in enumerate(codigos):
            for cb in codigos[i+1:]:
                ua = URI(D + "Residente/" + ca)
                ub = URI(D + "Residente/" + cb)
                sc = matriz_afinidad.at[ca, cb]
                pa = perfiles[ca]; pb = perfiles[cb]

                if sc > 0:
                    gi.add((ua, _COMPAT, ub)); gi.add((ub, _COMPAT, ua))
                    n_compat += 2
                    # comparteDieta si ambos tienen la misma dieta especial
                    da = pa["dieta"] - {"Ninguna"}; db = pb["dieta"] - {"Ninguna"}
                    if da and db and (da & db):
                        gi.add((ua, _DIETA, ub)); gi.add((ub, _DIETA, ua))
                        n_dieta += 2
                    # comparteAficion si mismo hobby
                    if pa.get("hobby") and pb.get("hobby") and pa["hobby"] == pb["hobby"]:
                        gi.add((ua, _AFICION, ub)); gi.add((ub, _AFICION, ua))
                        n_aficion += 2
                else:
                    gi.add((ua, _INCOMPAT, ub)); gi.add((ub, _INCOMPAT, ua))
                    n_incompat += 2

        ruta = directorio / "dataset_inferido.ttl"
        gi.serialize(destination=str(ruta), format="turtle")
        total_orig = len(grafo_original); total_new = len(gi)
        print(f"  ✓ {ruta.name}  "
              f"({total_orig:,} tripletas originales + {total_new - total_orig:,} inferidas = {total_new:,} total)")
        print(f"    → Nivel: {n_nivel}  Compat: {n_compat}  Incompat: {n_incompat}  "
              f"Dieta: {n_dieta}  Afición: {n_aficion}  Hobby: {n_hobby}  Religión: {n_relig}")

    def exportar_alertas_clinicas_y_menus(self, perfiles: dict, directorio: Path) -> None:
        directorio.mkdir(parents=True, exist_ok=True)

        # ── alertas ──────────────────────────────────────────────────────────────
        alertas = []
        for code, p in perfiles.items():
            def _alerta(prio, motivo, accion):
                alertas.append({
                    "Prioridad":            prio,
                    "Código":               code,
                    "Nombre":               p["nombre"],
                    "Nivel Convivencia":    NIVEL_SHORT[p["nivel"]],
                    "Alerta":               motivo,
                    "Acción Recomendada":   accion,
                    "Barthel":              p.get("barthel"),
                    "GDS":                  p.get("gds"),
                    "Goldberg":             p.get("goldberg"),
                    "MEC-Lobo":             p.get("meclobo"),
                    "Norton":               p.get("norton"),
                    "Tinetti":              p.get("tinetti"),
                    "Pfeiffer":             p.get("pfeiffer"),
                    "NPI":                  p.get("npi"),
                    "NECPAL":               p.get("necpal"),
                })

            npi     = p.get("npi")     or 0
            necpal  = p.get("necpal")  or 0
            barthel = p.get("barthel")
            norton  = p.get("norton")
            tinetti = p.get("tinetti")
            gds     = p.get("gds")     or 0
            meclobo = p.get("meclobo")
            goldberg= p.get("goldberg") or 0

            # P1 
            if npi >= 4:
                _alerta("P1 · Crítica", f"Trastorno conductual grave (NPI={npi})",
                        "Evaluación psiquiátrica urgente; vigilancia continua")
            if necpal >= 1:
                _alerta("P1 · Crítica", "NECPAL positivo — pronóstico de vida limitado",
                        "Plan de cuidados paliativos individualizado")
            if barthel is not None and barthel <= 20:
                _alerta("P1 · Crítica", f"Dependencia total (Barthel={barthel})",
                        "Cuidados de enfermería intensivos 24 h")

            # P2 
            if norton is not None and norton <= 12:
                _alerta("P2 · Alta", f"Riesgo alto/muy alto úlceras presión (Norton={norton})",
                        "Protocolo de cambios posturales; colchón antiescaras")
            if tinetti is not None and tinetti < 10:
                _alerta("P2 · Alta", f"Alto riesgo de caídas (Tinetti={tinetti})",
                        "Barras de apoyo, cama articulada, fisioterapia prioritaria")
            if gds >= 12:
                _alerta("P2 · Alta", f"Depresión grave (GDS={gds})",
                        "Derivación a psicología/psiquiatría; reevaluación mensual")
            if meclobo is not None and meclobo <= 10:
                _alerta("P2 · Alta", f"Deterioro cognitivo grave (MEC-Lobo={meclobo})",
                        "Estimulación cognitiva intensiva; supervisión constante")
            if p.get("usaPaniales"):
                _alerta("P2 · Alta", "Uso de pañales — incontinencia total",
                        "Protocolo de higiene íntima; registro de incontinencia")
            if p.get("usaSillaRuedas"):
                _alerta("P2 · Alta", "Dependencia de silla de ruedas",
                        "Verificar accesibilidad habitación y zonas comunes")

            # P3
            if goldberg >= 5:
                _alerta("P3 · Media", f"Ansiedad moderada-alta (Goldberg={goldberg})",
                        "Seguimiento psicológico; técnicas de relajación")
            elif goldberg >= 4:
                _alerta("P3 · Media", f"Ansiedad leve (Goldberg={goldberg})",
                        "Observación y apoyo social")
            if 5 <= gds <= 11:
                _alerta("P3 · Media", f"Depresión leve-moderada (GDS={gds})",
                        "Actividades de motivación; contacto familiar regular")
            if norton is not None and 13 <= norton <= 14:
                _alerta("P3 · Media", f"Riesgo medio úlceras presión (Norton={norton})",
                        "Cambios posturales periódicos; hidratación de piel")
            if tinetti is not None and 10 <= tinetti <= 18:
                _alerta("P3 · Media", f"Riesgo medio de caídas (Tinetti={tinetti})",
                        "Fisioterapia preventiva; calzado antideslizante")
            if p.get("perdidaMemoria"):
                _alerta("P3 · Media", "Pérdida de memoria declarada",
                        "Taller de estimulación cognitiva; supervisión en AVD")
            if p.get("necesitaAsistNoc"):
                _alerta("P3 · Media", "Necesita asistencia nocturna",
                        "Valorar turno de noche dedicado o sistema de llamada")

        # Ordenar: 
        prioridad_ord = {"P1 · Crítica": 0, "P2 · Alta": 1, "P3 · Media": 2}
        alertas.sort(key=lambda x: (prioridad_ord.get(x["Prioridad"], 9), x["Nombre"]))

        p1 = sum(1 for a in alertas if a["Prioridad"].startswith("P1"))
        p2 = sum(1 for a in alertas if a["Prioridad"].startswith("P2"))
        p3 = sum(1 for a in alertas if a["Prioridad"].startswith("P3"))

        # ── menus ─────────────────────────────────────────────────────────────────
        menus = []
        for code, p in perfiles.items():
            dietas = p["dieta"] - {"Ninguna"}
            menus_asig = [_DIET_MENU[d] for d in dietas if d in _DIET_MENU]
            if p.get("tieneDiabetes"):
                if "Menú Diabético" not in menus_asig:
                    menus_asig.append("Menú Diabético")
            col = p.get("colesterol") or 0
            if col >= 240:
                menus_asig.append("Bajo en Colesterol")

            barthel = p.get("barthel")
            necesita_ayuda = (
                barthel is not None and barthel <= 40
                or p.get("ayudaComer") is True
            )
            menus.append({
                "Código":               code,
                "Nombre":               p["nombre"],
                "Nivel Convivencia":    NIVEL_SHORT[p["nivel"]],
                "Menú Principal":       "Estándar" if not menus_asig else menus_asig[0],
                "Dietas Especiales":    ", ".join(menus_asig) or "Ninguna",
                "Diabetes":             "Sí" if p.get("tieneDiabetes") else "No",
                "Colesterol Alto":      "Sí" if col >= 240 else "No",
                "Triturado/Puré":       "Sí" if "Triturado/Puré" in menus_asig else "No",
                "Barthel":              barthel,
                "Necesita Ayuda Comer": "Sí" if necesita_ayuda else "No",
                "Consideraciones":      "; ".join(filter(None, [
                    "Dieta hipocalórica" if p.get("tieneObesidad") else "",
                    "Sin alcohol" if not p.get("consumeAlcohol") else "",
                    "Supervisión al comer" if necesita_ayuda else "",
                ])),
            })

        menus.sort(key=lambda x: x["Nombre"])
        especiales = sum(1 for m in menus if m["Dietas Especiales"] != "Ninguna")

        # ── Excel formateado: alertas_enfermeria.xlsx + menus_comedor.xlsx ──────────
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("  ⚠ openpyxl no disponible → Excel no generado (pip install openpyxl)")
            return

        def _fill(h):  return PatternFill("solid", fgColor=h)
        def _font(bold=False, color="1A1A1A", size=9, italic=False):
            return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
        def _align(h="left", wrap=False):
            return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
        _s = Side(style="thin", color="CCCCCC")
        _b = Border(left=_s, right=_s, top=_s, bottom=_s)

        # ── EXCEL 1: alertas_enfermeria.xlsx ─────────────────────────────────────
        wb_a = Workbook()
        ws1 = wb_a.active
        ws1.title = "Alertas Enfermería"
        ws1.sheet_view.showGridLines = False

        PRIO_BG = {"P1 · Crítica": "FFCDD2", "P2 · Alta": "FFE0B2", "P3 · Media": "FFF9C4"}
        PRIO_FG = {"P1 · Crítica": "B71C1C", "P2 · Alta": "BF360C", "P3 · Media": "E65100"}

        ncA = 15
        ws1.merge_cells(f"A1:{get_column_letter(ncA)}1")
        c = ws1["A1"]
        c.value = (f"ALERTAS CLÍNICAS DE ENFERMERÍA  ·  {len(alertas)} alertas  "
                   f"(P1 Crítica: {p1}  ·  P2 Alta: {p2}  ·  P3 Media: {p3})")
        c.fill = _fill("1A3A5C"); c.font = _font(bold=True, color="FFFFFF", size=12)
        c.alignment = _align("center"); ws1.row_dimensions[1].height = 26

        hdrs_a = ["Prioridad", "Código", "Nombre", "Nivel", "Alerta", "Acción Recomendada",
                  "Barthel", "GDS", "Goldberg", "MEC-Lobo", "Norton", "Tinetti",
                  "Pfeiffer", "NPI", "NECPAL"]
        for j, h in enumerate(hdrs_a, 1):
            c = ws1.cell(2, j, h)
            c.fill = _fill("2D6A9F"); c.font = _font(bold=True, color="FFFFFF")
            c.alignment = _align("center", wrap=True); c.border = _b
        ws1.row_dimensions[2].height = 32
        ws1.freeze_panes = "A3"

        for j, w in enumerate([15, 10, 16, 22, 52, 52, 8, 7, 9, 9, 8, 8, 9, 7, 8], 1):
            ws1.column_dimensions[get_column_letter(j)].width = w

        for i, a in enumerate(alertas, 3):
            prio = a["Prioridad"]
            bg = PRIO_BG.get(prio, "FFFFFF")
            fg = PRIO_FG.get(prio, "000000")
            for j, key in enumerate(hdrs_a, 1):
                val = a.get(key) if a.get(key) is not None else "—"
                c = ws1.cell(i, j, val)
                c.fill = _fill(bg)
                c.font = _font(bold=(j == 1), color=(fg if j == 1 else "1A1A1A"))
                c.alignment = _align(
                    "center" if j in (1, 2, 7, 8, 9, 10, 11, 12, 13, 14, 15) else "left",
                    wrap=(j in (5, 6)))
                c.border = _b
            txt_accion = a.get("Acción Recomendada", "")
            ws1.row_dimensions[i].height = 30 if len(txt_accion) > 55 else 16

        ruta_a = directorio / "alertas_enfermeria.xlsx"
        try:
            wb_a.save(str(ruta_a))
            print(f"  ✓ alertas_enfermeria.xlsx  ({len(alertas)} alertas: P1={p1} P2={p2} P3={p3})")
        except PermissionError:
            import time
            alt = directorio / f"alertas_enfermeria_{int(time.time())}.xlsx"
            wb_a.save(str(alt))
            print(f"  ⚠ alertas_enfermeria.xlsx estaba abierto → guardado como {alt.name}")

        # ── EXCEL 2: menus_comedor.xlsx ───────────────────────────────────────────
        wb_m = Workbook()
        ws2 = wb_m.active
        ws2.title = "Menús Comedor"
        ws2.sheet_view.showGridLines = False

        MENU_BG = {
            "Menú Diabético":     "E8F5E9",
            "Sin Gluten":         "FFF8E1",
            "Sin Lactosa":        "E3F2FD",
            "Hipocalórico":       "FCE4EC",
            "Hiposódico":         "F3E5F5",
            "Triturado/Puré":     "FBE9E7",
            "Vegetariano":        "DCEDC8",
            "Vegano":             "C5E1A5",
            "Bajo en Colesterol": "FFF3E0",
            "Blanda":             "EFEBE9",
        }

        ncM = 11
        ws2.merge_cells(f"A1:{get_column_letter(ncM)}1")
        c = ws2["A1"]
        c.value = (f"MENÚS DE COMEDOR ASIGNADOS  ·  {len(menus)} residentes  "
                   f"·  {especiales} con dieta especial")
        c.fill = _fill("1A3A5C"); c.font = _font(bold=True, color="FFFFFF", size=12)
        c.alignment = _align("center"); ws2.row_dimensions[1].height = 26

        hdrs_m = ["Código", "Nombre", "Nivel", "Menú Principal", "Dietas Especiales",
                  "Diabetes", "Colesterol", "Triturado", "Barthel", "Ayuda Comer",
                  "Consideraciones"]
        m_keys = ["Código", "Nombre", "Nivel Convivencia", "Menú Principal", "Dietas Especiales",
                  "Diabetes", "Colesterol Alto", "Triturado/Puré", "Barthel",
                  "Necesita Ayuda Comer", "Consideraciones"]
        for j, h in enumerate(hdrs_m, 1):
            c = ws2.cell(2, j, h)
            c.fill = _fill("2D6A9F"); c.font = _font(bold=True, color="FFFFFF")
            c.alignment = _align("center", wrap=True); c.border = _b
        ws2.row_dimensions[2].height = 28
        ws2.freeze_panes = "A3"

        for j, w in enumerate([10, 16, 22, 18, 32, 9, 11, 10, 8, 12, 32], 1):
            ws2.column_dimensions[get_column_letter(j)].width = w

        for i, m in enumerate(menus, 3):
            menu_p = m.get("Menú Principal", "Estándar")
            bg = MENU_BG.get(menu_p, "FFFFFF" if i % 2 == 0 else "F5F5F5")
            for j, key in enumerate(m_keys, 1):
                val = m.get(key) if m.get(key) is not None else "—"
                c = ws2.cell(i, j, val)
                c.fill = _fill(bg)
                c.font = _font(bold=(j == 4 and menu_p != "Estándar"))
                c.alignment = _align(
                    "center" if j in (1, 6, 7, 8, 9, 10) else "left",
                    wrap=(j in (5, 11)))
                c.border = _b
            ws2.row_dimensions[i].height = 15

        ruta_m = directorio / "menus_comedor.xlsx"
        try:
            wb_m.save(str(ruta_m))
            print(f"  ✓ menus_comedor.xlsx  ({len(menus)} residentes, {especiales} con dieta especial)")
        except PermissionError:
            import time
            alt = directorio / f"menus_comedor_{int(time.time())}.xlsx"
            wb_m.save(str(alt))
            print(f"  ⚠ menus_comedor.xlsx estaba abierto → guardado como {alt.name}")


    def exportar_ranking_parejas(self, parejas: list, perfiles: dict, directorio: Path) -> None:
        """Genera ranking_parejas.xlsx con todas las parejas no vetadas ordenadas por afinidad."""
        directorio.mkdir(parents=True, exist_ok=True)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("  ⚠ openpyxl no disponible → ranking_parejas no generado")
            return

        def _fill(h):  return PatternFill("solid", fgColor=h)
        def _font(bold=False, color="1A1A1A", size=9):
            return Font(bold=bold, color=color, size=size, name="Calibri")
        def _align(h="center", wrap=False):
            return Alignment(horizontal=h, vertical="center", wrap_text=wrap)
        _s = Side(style="thin", color="CCCCCC")
        _b = Border(left=_s, right=_s, top=_s, bottom=_s)

        def _score_bg(pct):
            if pct >= 70: return "C8E6C9"   # verde
            if pct >= 50: return "DCEDC8"   # verde claro
            if pct >= 30: return "FFF9C4"   # amarillo
            return "FFECB3"                 # naranja claro

        no_vet = [p for p in parejas if p["pct"] > 0]

        wb = Workbook()
        ws = wb.active
        ws.title = "Ranking Parejas"
        ws.sheet_view.showGridLines = False

        ncR = 9
        ws.merge_cells(f"A1:{get_column_letter(ncR)}1")
        c = ws["A1"]
        c.value = (f"RANKING DE PAREJAS COMPATIBLES  ·  {len(no_vet):,} parejas no vetadas  "
                   f"(de {len(parejas):,} evaluadas en total)")
        c.fill = _fill("1A3A5C"); c.font = _font(bold=True, color="FFFFFF", size=12)
        c.alignment = _align("center"); ws.row_dimensions[1].height = 26

        hdrs = ["Pos.", "Código A", "Nombre A", "Nivel A",
                "Código B", "Nombre B", "Nivel B",
                "Afinidad %", "Principales razones (+)"]
        for j, h in enumerate(hdrs, 1):
            c = ws.cell(2, j, h)
            c.fill = _fill("2D6A9F"); c.font = _font(bold=True, color="FFFFFF")
            c.alignment = _align("center", wrap=True); c.border = _b
        ws.row_dimensions[2].height = 30
        ws.freeze_panes = "A3"

        for j, w in enumerate([6, 11, 16, 24, 11, 16, 24, 12, 58], 1):
            ws.column_dimensions[get_column_letter(j)].width = w

        for i, par in enumerate(no_vet, 3):
            bg = _score_bg(par["pct"])
            razones = "  ·  ".join(
                txt for txt, pts in sorted(par["detalle"], key=lambda x: -x[1])[:4]
                if pts > 0
            )
            nA = NIVEL_SHORT.get(par["nivel_A"], "")
            nB = NIVEL_SHORT.get(par["nivel_B"], "")
            vals = [i - 2, par["A"], par["nombre_A"], nA,
                    par["B"], par["nombre_B"], nB, par["pct"], razones]
            for j, v in enumerate(vals, 1):
                c = ws.cell(i, j, v)
                c.fill = _fill(bg)
                c.font = _font(
                    bold=(j == 8),
                    color=("2E7D32" if par["pct"] >= 70 else "1A1A1A"))
                c.alignment = _align("center" if j != 9 else "left", wrap=(j == 9))
                c.border = _b
            ws.row_dimensions[i].height = 15

        ruta_x = directorio / "ranking_parejas.xlsx"
        try:
            wb.save(str(ruta_x))
            print(f"  ✓ ranking_parejas.xlsx  ({len(no_vet):,} parejas no vetadas)")
        except PermissionError:
            import time
            alt = directorio / f"ranking_parejas_{int(time.time())}.xlsx"
            wb.save(str(alt))
            print(f"  ⚠ ranking_parejas.xlsx estaba abierto → guardado como {alt.name}")


# CLASE: TFG_SistemaCoHousing

class TFG_SistemaCoHousing:
    """Coordinador principal del sistema: consola + orquestación de las 3 subclases."""

    def __init__(self):
        self.cargador   = TFG_CargadorOntologia()
        self.motor      = TFG_MotorCompatibilidad()
        self.exportador = TFG_ExportadorResultados()

    @staticmethod
    def _limpiar_resultados(directorio: Path) -> None:
        """Elimina archivos no-Excel y xlsx con timestamp de ejecuciones anteriores."""
        import re
        if not directorio.exists():
            return
        eliminados = []
        for f in directorio.iterdir():
            if not f.is_file():
                continue
            # Borrar solo .csv (los .ttl y .xlsx se gestionan aparte)
            if f.suffix.lower() == ".csv":
                f.unlink()
                eliminados.append(f.name)
                continue
            # Borrar xlsx con timestamp (_XXXXXXXXXX.xlsx)
            if re.search(r"_\d{9,}\.xlsx$", f.name):
                f.unlink()
                eliminados.append(f.name)
                continue
            # Borrar alertas_menus.xlsx si quedó de versiones anteriores
            if f.name == "alertas_menus.xlsx":
                f.unlink()
                eliminados.append(f.name)
        if eliminados:
            print(f"  🗑  Limpieza: {len(eliminados)} archivo(s) antiguo(s) eliminado(s)")

    def ejecutar(self, args):
        print("Iniciando algoritmo...")

        self._limpiar_resultados(RESULTADOS)

        grafo = self.cargador.cargar_grafo_ttl(TTL_FILE)
        print("Ontologia cargada correctamente.")

        datos_excel_xls = self.cargador.cargar_datos_excel_xls()
        perfiles = self.cargador.cargar_todos_perfiles(grafo, datos_excel_xls)
        print(f"Perfiles cargados: {len(perfiles)} residentes.")

        if args.residente and args.residente not in perfiles:
            print(f"Error: el codigo '{args.residente}' no existe.")
            print(f"Algunos codigos disponibles: {', '.join(sorted(perfiles)[:10])} ...")
            return

        matriz_afinidad, parejas, ranking = self.motor.calcular_matriz_afinidad_completa(perfiles)
        print(f"Compatibilidad calculada: {len(parejas):,} parejas evaluadas.")

        if not args.sin_excel:
            self.exportador.exportar_excel_profesional(perfiles, matriz_afinidad, parejas, ranking, RESULTADOS)

        self.exportador.exportar_grafo_semantico_enriquecido(grafo, perfiles, matriz_afinidad, RESULTADOS)
        self.exportador.exportar_alertas_clinicas_y_menus(perfiles, RESULTADOS)
        self.exportador.exportar_ranking_parejas(parejas, perfiles, RESULTADOS)

        print("Todo correcto. Resultados guardados en la carpeta 'resultados/'.")


# MAIN
def main() -> None:
    parser = argparse.ArgumentParser(
        description="Calcula % de compatibilidad entre residentes co-housing")
    parser.add_argument("residente", nargs="?", default=None,
                        help="Código de residente para filtrar consola (ej. BMW7812)")
    parser.add_argument("--top", type=int, default=10,
                        help="Cuántos candidatos mostrar por residente (default: 10)")
    parser.add_argument("--sin-excel", action="store_true",
                        help="No generar el Excel")
    args = parser.parse_args()
    TFG_SistemaCoHousing().ejecutar(args)


if __name__ == "__main__":
    main()
